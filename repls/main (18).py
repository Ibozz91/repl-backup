def decsum(arr):
  x=Decimal(0)
  for i in arr:
    x=Decimal(x)+Decimal(i)
  return x
#File must be named "votes.xlsx"
from openpyxl import load_workbook
from decimal import *
workbook=load_workbook(filename="votes.xlsx")
sheet=workbook.active
candidates=[]
qw=1
for i in sheet["1"]:
  if i in candidates:
    candidates.append(i.value+qw)
    qw+=1
  else:
    candidates.append(i.value)
votes=[]
print("All candidates: "+str(candidates))
i=0
maxs=0
for g in sheet.iter_rows(min_row=2, values_only=True):
  votes.append([])
  for h in g:
    if int(h)>maxs:
      maxs=int(h)
    votes[i].append(int(h))
  i+=1
for z in range(0,len(votes)):
  for c in range(0,len(votes[0])):
    votes[z][c]=Decimal(votes[z][c])/Decimal(maxs)
seats=int(input("How many seats to fill?\n"))
harequota=Decimal(len(votes))/Decimal(seats)
moneypervoter=[]
for jkl in range(0, len(votes)):
  moneypervoter.append(Decimal(1))
print("Quota = $"+str(harequota))
elected=[]
while len(elected)<seats:
  print("ROUND "+str(len(elected)+1))
  p_affordability=[]
  for i in range(0, len(candidates)):
    if candidates[i] in elected:
      p_affordability.append(False)
    else:
      pq=[]
      for y in range(0, len(votes)):
        pq.append((votes[y][i]>0)*Decimal(moneypervoter[y]))
      if decsum(pq)>=harequota:
        p_affordability.append(True)
      else:
        p_affordability.append(False)
  if True in p_affordability:
    highestp=[]
    totalspv=[]
    for mnb in range(0, len(candidates)):
      if p_affordability[mnb] and (not candidates[mnb] in elected):
        spendpervoter=[]
        for plk in votes:
          spendpervoter.append(Decimal(0))
        pmin=Decimal(0)
        pmax=maxs
        while pmax-pmin>0.000001:
          spendpervoter=[]
          for plk in votes:
            spendpervoter.append(Decimal(0))
          ptest=(pmax+pmin)/2
          for asdf in range(0, len(votes)):
            spendpervoter[asdf]=min([votes[asdf][mnb]*ptest, moneypervoter[asdf]])
          if decsum(spendpervoter)>harequota:
            pmax=ptest
          elif decsum(spendpervoter)<harequota:
            pmin=ptest
          else:
            pmax=ptest
            pmin=ptest
        totalspv.append(spendpervoter)
        highestp.append((pmax+pmin)/2)
      else:
        lol=[]
        for yuio in votes:
          lol.append(maxs+1)
        totalspv.append(lol)
        highestp.append(maxs+1)
    jj=2
    jjn=-1
    for jjj in range(0, len(candidates)):
      if highestp[jjj]<=maxs:
        print(candidates[jjj]+" can be afforded at p=$"+str(highestp[jjj]))
      elif candidates[jjj] in elected:
        print(candidates[jjj]+" has already been elected.")
      else:
        print(candidates[jjj]+" is not affordable by voters for the price of a quota.")
      if highestp[jjj]<jj:
        jj=highestp[jjj]
        jjn=jjj
    print(candidates[jjn]+" is chosen.")
    elected.append(candidates[jjn])
    for outofnames in range(0, len(votes)):
      moneypervoter[outofnames]-=totalspv[jjn][outofnames]
  else:
    print("No candidate is affordable.")
    jjn=-1
    jj=-1
    for jjj in range(0, len(candidates)):
      plmn=Decimal(0)
      for k in range(0, len(votes)):
        plmn+=votes[k][jjj]*moneypervoter[k]
      if candidates[jjj] in elected:
        print(candidates[jjj]+" has already been elected.")
      else:
        print(candidates[jjj]+" would have voters spend $"+str(plmn)+".")
      if plmn>jj and not candidates[jjj] in elected:
        jj=plmn
        jjn=jjj
    print(candidates[jjn]+" is chosen.")
    elected.append(candidates[jjn])
    for outofnames in range(0, len(votes)):
      moneypervoter[outofnames]-=votes[outofnames][jjn]*moneypervoter[outofnames]
print("Elected: "+str(elected))